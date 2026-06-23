import pytest
from django.core.management import call_command
from rest_framework.test import APIClient

pytestmark = pytest.mark.django_db


@pytest.fixture
def seeded(db):
    call_command("load_services_fixture")


@pytest.fixture
def client():
    return APIClient()


@pytest.fixture
def auth_client(seeded):
    """APIClient authenticated as the chief admin (for protected endpoints)."""
    c = APIClient()
    tok = c.post(
        "/api/auth/login", {"username": "admin", "password": "admin"}, format="json"
    ).json()["token"]
    c.credentials(HTTP_AUTHORIZATION=f"Bearer {tok}")
    return c


def test_login_returns_contract_shape(seeded, client):
    r = client.post("/api/auth/login", {"username": "admin", "password": "admin"}, format="json")
    assert r.status_code == 200
    body = r.json()
    assert {"token", "user_id", "username", "role", "counter_id", "expires_at"} <= set(body)
    assert body["role"] == "admin"


def test_login_rejects_bad_credentials(seeded, client):
    r = client.post("/api/auth/login", {"username": "admin", "password": "nope"}, format="json")
    assert r.status_code == 401


def test_categories_and_services(seeded, client):
    cats = client.get("/api/categories").json()
    assert len(cats) == 9
    assert set(cats[0]) == {
        "id", "hall_id", "code", "name_kaa", "name_ru", "name_uz", "name_en",
        "color", "order",
    }

    svcs = client.get("/api/services?category_id=1").json()
    assert all(s["category_id"] == 1 for s in svcs)
    assert {"id", "category_id", "name_kaa", "name_ru", "name_uz", "name_en",
            "sla_days", "delivery_type", "requires_visit", "is_active", "is_popular"} == set(svcs[0])


def test_category_create_and_delete(auth_client):
    client = auth_client
    before = len(client.get("/api/categories").json())
    r = client.post("/api/categories", {
        "code": "Z", "name_kaa": "Test", "name_ru": "Тест", "color": "#000000", "order": 99,
    }, format="json")
    assert r.status_code == 201
    cid = r.json()["id"]
    assert len(client.get("/api/categories").json()) == before + 1
    d = client.delete(f"/api/categories/{cid}")
    assert d.status_code == 204
    assert len(client.get("/api/categories").json()) == before


def test_category_create_requires_auth(seeded, client):
    r = client.post("/api/categories", {
        "code": "Z", "name_kaa": "T", "name_ru": "Т", "color": "#000000", "order": 99,
    }, format="json")
    assert r.status_code in (401, 403)


def test_service_create_update_delete(auth_client):
    client = auth_client
    r = client.post("/api/services", {
        "category_id": 1, "name_kaa": "Jańa xızmet", "name_ru": "Новая услуга",
        "sla_days": 1, "delivery_type": "electron", "requires_visit": True,
        "is_active": True, "is_popular": False,
    }, format="json")
    assert r.status_code == 201
    body = r.json()
    assert body["category_id"] == 1
    sid = body["id"]
    patched = client.patch(f"/api/services/{sid}", {"is_popular": True}, format="json")
    assert patched.json()["is_popular"] is True
    assert client.delete(f"/api/services/{sid}").status_code == 204


def test_counters_serialize_service_ids(seeded, client):
    counters = client.get("/api/counters").json()
    c1 = next(c for c in counters if c["id"] == 1)
    assert isinstance(c1["service_ids"], list) and len(c1["service_ids"]) == 9


def test_kiosk_ticket_create_is_idempotent(seeded, client):
    payload = {"category_id": 1, "service_id": 1, "idempotency_key": "abc"}
    r1 = client.post("/api/tickets", payload, format="json")
    assert r1.status_code == 201
    t1 = r1.json()
    assert t1["number"].startswith("A")
    assert t1["status"] == "waiting"
    r2 = client.post("/api/tickets", payload, format="json")
    assert r2.json()["id"] == t1["id"]


def test_operator_flow_call_finish(seeded, client):
    # queue for counter 1 (serves services incl. 1)
    client.post("/api/tickets", {"category_id": 1, "service_id": 1, "idempotency_key": "k1"}, format="json")

    q = client.get("/api/queue?counter_id=1").json()
    assert len(q) >= 1

    called = client.post("/api/tickets/call-next", {"counter_id": 1, "operator_id": 2}, format="json")
    assert called.status_code == 200
    ticket = called.json()
    assert ticket["status"] == "called"
    assert ticket["counter_id"] == 1

    current = client.get("/api/tickets/current?counter_id=1").json()
    assert current["id"] == ticket["id"]

    fin = client.post(f"/api/tickets/{ticket['id']}/finish")
    assert fin.json()["status"] == "served"

    # current clears
    assert client.get("/api/tickets/current?counter_id=1").json() is None


def test_status_change_bumps_updated_at(seeded, client):
    """Регресс: call/finish/skip ДОЛЖНЫ двигать updated_at (auto_now). Иначе
    local→cloud sync (watermark по updated_at) не видит смену статуса и
    обслуженные талоны застревают как 'waiting' в облаке."""
    from queue_app.models import Ticket

    tid = client.post(
        "/api/tickets",
        {"category_id": 1, "service_id": 1, "idempotency_key": "upd-bump"},
        format="json",
    ).json()["id"]
    u_created = Ticket.objects.get(id=tid).updated_at

    client.post("/api/tickets/call-next", {"counter_id": 1, "operator_id": 2}, format="json")
    u_called = Ticket.objects.get(id=tid).updated_at
    assert u_called > u_created, "call-next должен обновлять updated_at"

    client.post(f"/api/tickets/{tid}/finish")
    u_finished = Ticket.objects.get(id=tid).updated_at
    assert u_finished > u_called, "finish должен обновлять updated_at"


def test_display_active_after_call(seeded, client):
    client.post("/api/tickets", {"category_id": 1, "service_id": 1, "idempotency_key": "k1"}, format="json")
    client.post("/api/tickets/call-next", {"counter_id": 1, "operator_id": 2}, format="json")
    calls = client.get("/api/display/active").json()
    assert len(calls) == 1
    assert set(calls[0]) == {
        "id", "number", "category_id", "counter_id",
        "counter_number", "counter_name", "called_at", "status",
    }
    assert calls[0]["counter_number"] == "1"


def test_operator_session_lifecycle(seeded, client):
    r = client.post("/api/operator-sessions", {"user_id": 2, "counter_id": 1}, format="json")
    assert r.status_code == 201
    sid = r.json()["id"]
    patched = client.patch(f"/api/operator-sessions/{sid}", {"status": "break"}, format="json")
    assert patched.json()["status"] == "break"


def test_dashboard_shape(seeded, client):
    body = client.get("/api/dashboard").json()
    assert set(body) == {"metrics", "hourly", "recent"}
    assert set(body["metrics"]) == {"ticketsToday", "avgWaitMinutes", "activeCounters", "served"}
    assert body["metrics"]["activeCounters"] == 5


def test_display_settings_get_and_patch(auth_client):
    client = auth_client
    r = client.get("/api/display/settings")
    assert r.status_code == 200
    assert "youtube_url" in r.json()

    p = client.patch(
        "/api/display/settings",
        {"youtube_url": "https://youtu.be/abc123"},
        format="json",
    )
    assert p.status_code == 200
    assert p.json()["youtube_url"] == "https://youtu.be/abc123"
    # persisted (singleton)
    assert client.get("/api/display/settings").json()["youtube_url"] == "https://youtu.be/abc123"


def test_display_board_lists_active_windows(seeded, client):
    board = client.get("/api/display/board").json()
    assert len(board) >= 1
    row = board[0]
    assert set(row) == {"counter_id", "counter_number", "counter_name", "current"}
    # all windows idle right after seeding
    assert all(w["current"] is None for w in board)


def test_display_waiting_lists_uncalled_tickets(seeded, client):
    # seeded fixture pre-populates the waiting queue
    waiting = client.get("/api/display/waiting").json()
    assert isinstance(waiting, list)
    if waiting:
        assert set(waiting[0]) == {"id", "number", "category_id"}


def test_halls_endpoint(seeded, client):
    halls = client.get("/api/halls").json()
    assert len(halls) == 2
    assert {h["code"] for h in halls} == {"1", "2"}


def test_categories_are_global_across_halls(seeded, client):
    """Каталог общий для всех залов: hall-фильтр чтения убран (маршрут талона —
    по услугам окна, не по залу). Категории видны независимо от зала."""
    assert len(client.get("/api/categories").json()) == 9
    # ?hall_id больше НЕ сужает выборку — каталог одинаков для любого зала
    assert len(client.get("/api/categories?hall_id=2").json()) == 9


def test_display_active_scoped_by_hall(seeded, client):
    # Issue + call a ticket in hall 1, then check hall scoping on the board.
    t = client.post("/api/tickets", {
        "category_id": 1, "service_id": 1, "idempotency_key": "hall-test-1",
    }, format="json").json()
    assert t["hall_id"] == 1
    # counter 1 is in hall 1
    called = client.post("/api/tickets/call-next", {"counter_id": 1}, format="json")
    assert called.status_code == 200
    assert len(client.get("/api/display/active?hall_id=1").json()) >= 1
    assert len(client.get("/api/display/active?hall_id=2").json()) == 0


def test_audit_logs_call_and_lists(auth_client):
    client = auth_client
    client.post("/api/tickets", {"category_id": 1, "service_id": 1, "idempotency_key": "au1"}, format="json")
    client.post("/api/tickets/call-next", {"counter_id": 1, "operator_id": 2}, format="json")
    logs = client.get("/api/audit").json()
    assert any(l["action"] == "ticket.called" for l in logs)
    called = next(l for l in logs if l["action"] == "ticket.called")
    assert set(called) == {"id", "actor_id", "actor_label", "action", "target", "meta", "created_at"}
    # filter by action
    only = client.get("/api/audit?action=ticket.called").json()
    assert all(l["action"] == "ticket.called" for l in only)


def test_ticket_recall(auth_client):
    client = auth_client
    client.post("/api/tickets", {"category_id": 1, "service_id": 1, "idempotency_key": "rc1"}, format="json")
    called = client.post("/api/tickets/call-next", {"counter_id": 1}, format="json").json()
    r = client.post(f"/api/tickets/{called['id']}/recall")
    assert r.status_code == 200
    logs = client.get("/api/audit?action=ticket.recalled").json()
    assert any(l["target"] == called["number"] for l in logs)


def test_stats_and_export(seeded, client):
    client.post("/api/tickets", {"category_id": 1, "service_id": 1, "idempotency_key": "st1"}, format="json")
    called = client.post("/api/tickets/call-next", {"counter_id": 1}, format="json").json()
    client.post(f"/api/tickets/{called['id']}/finish")
    s = client.get("/api/stats").json()
    assert s["served"] >= 1
    assert {"issued", "served", "skipped", "avg_wait_minutes",
            "avg_service_minutes", "peak_hour", "hourly"} <= set(s)
    r = client.get("/api/stats/export")
    assert r.status_code == 200
    # настоящий .xlsx (а не CSV)
    assert "spreadsheetml" in r["Content-Type"]
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(r.content))
    # три листа: талоны + сводка по операторам + услуги
    assert wb.sheetnames == ["Talonlar", "Operatorlar", "Xizmatlar"]
    ws = wb["Talonlar"]
    headers = [c.value for c in ws[1]]
    assert headers[:8] == [
        "Raqam", "Sana", "Zal", "Kategoriya", "Xizmat", "Oyna", "Operator", "Holat",
    ]
    assert ws.auto_filter.ref  # автофильтр (кнопки сортировки) включён
    # лист «Operatorlar» — сводка
    ws2 = wb["Operatorlar"]
    assert [c.value for c in ws2[1]][:3] == [
        "Operator", "Qabul qilingan (kishi)", "Jami xizmat (daq)",
    ]
    assert ws2.auto_filter.ref
    # лист «Xizmatlar» — популярность услуг
    ws3 = wb["Xizmatlar"]
    assert [c.value for c in ws3[1]][:3] == [
        "Xizmat", "Kategoriya", "So'ralgan (marta)",
    ]
    assert ws3.auto_filter.ref


def test_sync_catalog_snapshot(seeded, client):
    snap = client.get("/api/sync/catalog").json()
    assert {"halls", "categories", "services", "counters", "users", "settings"} <= set(snap)
    assert len(snap["categories"]) == 9
    assert len(snap["halls"]) == 2


def test_sync_events_ingest_idempotent(seeded, client):
    import uuid
    tid = str(uuid.uuid4())
    payload = {"tickets": [{
        "id": tid, "number": "A-999", "hall_id": 1, "category_id": 1,
        "status": "served", "created_at": "2026-06-04T10:00:00Z",
    }]}
    r = client.post("/api/sync/events", payload, format="json")
    assert r.status_code == 200 and r.json()["tickets"] == 1
    # re-push is idempotent (upsert by id)
    r2 = client.post("/api/sync/events", payload, format="json")
    assert r2.json()["tickets"] == 1
    from queue_app.models import Ticket
    assert Ticket.objects.filter(id=tid, number="A-999", status="served").count() == 1


def test_hall_admin_scoping(seeded, client):
    from accounts.models import User
    u = User.objects.create(username="ha2", role="hall_admin", hall_id=2, is_active=True)
    u.set_password("pw")
    u.save()
    tok = client.post("/api/auth/login", {"username": "ha2", "password": "pw"}, format="json").json()["token"]
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {tok}")
    # каталог ОБЩИЙ — hall_admin видит все категории (hall-фильтр чтения убран)
    assert len(client.get("/api/categories").json()) == 9
    # но создание всё ещё форсится в его зал (perform_create не трогали)
    r = client.post("/api/categories", {
        "code": "X", "name_kaa": "x", "name_ru": "х", "color": "#000000", "order": 1,
    }, format="json")
    assert r.status_code == 201
    assert r.json()["hall_id"] == 2
    assert len(client.get("/api/categories").json()) == 10


def test_hall_reset_clears_waiting(auth_client):
    client = auth_client
    client.post("/api/tickets", {"category_id": 1, "service_id": 1, "idempotency_key": "rs1"}, format="json")
    assert len(client.get("/api/display/waiting?hall_id=1").json()) > 0
    r = client.post("/api/halls/1/reset")
    assert r.status_code == 200 and r.json()["ok"] is True
    assert client.get("/api/display/waiting?hall_id=1").json() == []
    # reset is audited
    logs = client.get("/api/audit?action=queue.reset").json()
    assert any(l["target"] == "hall:1" for l in logs)


def test_catalog_changes_audited(auth_client):
    client = auth_client
    r = client.post("/api/categories", {
        "code": "Q", "name_kaa": "q", "name_ru": "к", "color": "#000000", "order": 50,
    }, format="json")
    assert r.status_code == 201
    logs = client.get("/api/audit?action=category.created").json()
    assert any(str(l["target"]) == str(r.json()["id"]) for l in logs)


# ---------- password change (self-service + chief reset) ----------
def test_self_change_password(auth_client):
    client = auth_client
    r = client.post("/api/auth/change-password", {
        "old_password": "admin", "new_password": "S3cret-pw!",
    }, format="json")
    assert r.status_code == 204
    # old password no longer works, new one does
    assert client.post("/api/auth/login", {"username": "admin", "password": "admin"}, format="json").status_code == 401
    assert client.post("/api/auth/login", {"username": "admin", "password": "S3cret-pw!"}, format="json").status_code == 200
    # the change is audited
    assert client.get("/api/audit?action=auth.password_changed").json()


def test_change_password_rejects_wrong_old(auth_client):
    r = auth_client.post("/api/auth/change-password", {
        "old_password": "WRONG", "new_password": "S3cret-pw!",
    }, format="json")
    assert r.status_code == 400


def test_change_password_requires_auth(seeded, client):
    r = client.post("/api/auth/change-password", {
        "old_password": "admin", "new_password": "S3cret-pw!",
    }, format="json")
    assert r.status_code in (401, 403)


def test_hall_admin_create_and_user_scoping(seeded, client):
    """Chief creates a hall_admin for hall 2; that head sees/manages only their
    hall's staff and can't mint admins."""
    chief = APIClient()
    tok = chief.post("/api/auth/login", {"username": "admin", "password": "admin"}, format="json").json()["token"]
    chief.credentials(HTTP_AUTHORIZATION=f"Bearer {tok}")
    # create the head of hall 2
    r = chief.post("/api/users", {
        "username": "head2", "name": "Head", "role": "hall_admin", "hall_id": 2, "password": "hp",
    }, format="json")
    assert r.status_code == 201, r.content
    assert r.json()["hall_id"] == 2 and r.json()["role"] == "hall_admin"

    # head2 logs in → login carries their hall
    ha = APIClient()
    login = ha.post("/api/auth/login", {"username": "head2", "password": "hp"}, format="json").json()
    assert login["hall_id"] == 2
    ha.credentials(HTTP_AUTHORIZATION=f"Bearer {login['token']}")

    # sees only hall-2 staff (themselves; the seeded operators are hall 1 / unset)
    users = ha.get("/api/users").json()
    assert users and all(u["hall_id"] == 2 for u in users)

    # creating an operator is forced into hall 2
    op = ha.post("/api/users", {"username": "op_h2", "name": "Op", "role": "operator", "password": "x"}, format="json")
    assert op.status_code == 201
    assert op.json()["hall_id"] == 2

    # a hall_admin cannot mint an admin/chief
    bad = ha.post("/api/users", {"username": "hacker", "name": "H", "role": "admin", "password": "x"}, format="json")
    assert bad.status_code == 403


def test_chief_resets_operator_password(auth_client):
    client = auth_client
    created = client.post("/api/users", {
        "username": "op_reset", "name": "Op", "role": "operator",
    }, format="json")
    assert created.status_code == 201
    uid = created.json()["id"]
    # chief sets a brand-new password via PATCH
    r = client.patch(f"/api/users/{uid}", {"password": "Fresh-pw-9!"}, format="json")
    assert r.status_code == 200
    assert "password" not in r.json()  # never echoed back
    assert client.post("/api/auth/login", {"username": "op_reset", "password": "Fresh-pw-9!"}, format="json").status_code == 200
    assert client.get("/api/audit?action=user.password_reset").json()


# ---------- work schedule (recurring shifts) ----------
def _admin_id():
    from accounts.models import User
    return User.objects.get(username="admin").id


def test_schedule_crud_and_validation(auth_client):
    client = auth_client
    uid = _admin_id()
    # create a Monday morning shift on counter 1
    r = client.post("/api/schedule", {
        "user_id": uid, "counter_id": 1, "weekday": 0,
        "start_time": "08:00", "end_time": "12:00",
    }, format="json")
    assert r.status_code == 201, r.content
    sid = r.json()["id"]
    assert r.json()["hall_id"] == 1  # mirrored from counter's hall
    assert r.json()["weekday_label"] == "Понедельник"
    # end before start is rejected
    bad = client.post("/api/schedule", {
        "user_id": uid, "counter_id": 1, "weekday": 1,
        "start_time": "12:00", "end_time": "08:00",
    }, format="json")
    assert bad.status_code == 400
    # list + filter by weekday
    assert len(client.get("/api/schedule?weekday=0").json()) == 1
    assert client.get("/api/schedule?weekday=3").json() == []
    # delete
    assert client.delete(f"/api/schedule/{sid}").status_code == 204
    assert client.get("/api/schedule").json() == []


def test_schedule_current_matches_now(auth_client):
    from django.utils import timezone
    client = auth_client
    uid = _admin_id()
    now = timezone.localtime()
    # an all-day shift for today's weekday must show up in "current"
    client.post("/api/schedule", {
        "user_id": uid, "counter_id": 1, "weekday": now.weekday(),
        "start_time": "00:00", "end_time": "23:59",
    }, format="json")
    # a shift for a different weekday must NOT
    client.post("/api/schedule", {
        "user_id": uid, "counter_id": 1, "weekday": (now.weekday() + 1) % 7,
        "start_time": "00:00", "end_time": "23:59",
    }, format="json")
    current = client.get("/api/schedule/current").json()
    assert len(current) == 1
    assert current[0]["user_id"] == uid


def test_schedule_hall_admin_scoped(seeded, client):
    """A hall-2 admin can't see or schedule onto a hall-1 counter."""
    from accounts.models import User
    u = User.objects.create(username="ha_sch", role="hall_admin", hall_id=2, is_active=True)
    u.set_password("pw"); u.save()
    tok = client.post("/api/auth/login", {"username": "ha_sch", "password": "pw"}, format="json").json()["token"]
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {tok}")
    # counter 1 is in hall 1 → forbidden
    r = client.post("/api/schedule", {
        "user_id": u.id, "counter_id": 1, "weekday": 0,
        "start_time": "08:00", "end_time": "12:00",
    }, format="json")
    assert r.status_code == 403
    assert client.get("/api/schedule").json() == []
