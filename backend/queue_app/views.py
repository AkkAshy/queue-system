from django.db import transaction
from django.db.models import Avg, F
from django.http import JsonResponse
from django.utils import timezone
from rest_framework import generics
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import User
from accounts.permissions import (
    HasSyncToken,
    IsCatalogManager,
    IsChief,
    IsChiefOrReadOnly,
    IsStaff,
    hall_scope_ok,
    scope_to_hall,
)

from catalog.models import Service, ServiceCategory

from . import audit, realtime, services, sync
from .audit import AuditCRUDMixin
from .models import (
    AuditLog,
    Counter,
    DisplaySettings,
    OperatorSession,
    Ticket,
    TicketStatus,
    VoiceClip,
    WorkSchedule,
)
from .serializers import (
    AuditLogSerializer,
    CounterSerializer,
    CreateTicketSerializer,
    DisplayBoardCounterSerializer,
    DisplayCallSerializer,
    DisplaySettingsSerializer,
    DisplayWaitingSerializer,
    OperatorSessionSerializer,
    TicketSerializer,
    VoiceClipSerializer,
    WorkScheduleSerializer,
)


# ---------- counters ----------
class CounterListCreateView(AuditCRUDMixin, generics.ListCreateAPIView):
    audit_entity = "counter"
    serializer_class = CounterSerializer
    pagination_class = None
    permission_classes = [IsCatalogManager]

    def get_queryset(self):
        return scope_to_hall(Counter.objects.all(), self.request)

    def perform_create(self, serializer):
        u = self.request.user
        if getattr(u, "is_hall_admin", False) and u.hall_id:
            obj = serializer.save(hall_id=u.hall_id)
        else:
            obj = serializer.save()
        audit.log(self.request, "counter.created", target=obj.id)


class CounterDetailView(AuditCRUDMixin, generics.RetrieveUpdateDestroyAPIView):
    audit_entity = "counter"
    serializer_class = CounterSerializer
    permission_classes = [IsCatalogManager]

    def get_queryset(self):
        return scope_to_hall(Counter.objects.all(), self.request)


class HallResetView(APIView):
    """Clear a hall's queue (chief, or that hall's admin). TZ §4.4."""

    permission_classes = [IsStaff]

    def post(self, request, pk):
        if not hall_scope_ok(request.user, pk):
            return Response({"error": "forbidden"}, status=403)
        n = services.reset_queue(pk)
        audit.log(request, "queue.reset", target=f"hall:{pk}", cancelled=n)
        realtime.broadcast(
            [realtime.DISPLAY, realtime.OPERATORS, realtime.ADMIN], "queue.reset"
        )
        return Response({"ok": True, "cancelled": n})


# ---------- work schedule (shifts) ----------
class ScheduleListCreateView(AuditCRUDMixin, generics.ListCreateAPIView):
    """Recurring shifts. Chief sees/edits all; a hall_admin is scoped to their
    own hall (queryset filter + auto hall on create). TZ §4.2 (смены).

    POST assigns shifts in bulk across the operators × weekdays matrix: the
    admin picks days, then the operators who work them — each operator's window
    comes from their profile (``User.counter``), so the form never picks a
    counter. Shape::

        {"weekdays": [0, 1, 2], "user_ids": [5, 6],
         "start_time": "09:00", "end_time": "18:00", "is_active": true}

    Upserts per ``(user, counter, weekday)``: an existing day is updated (time/
    active), a missing one is created — so editing a shift's time never spawns a
    duplicate. The legacy single-row shape (``user_id``/``weekday``/
    ``counter_id``) is still accepted. Returns ``{created, updated, no_counter}``
    where ``no_counter`` lists operators skipped for lacking a window."""

    audit_entity = "schedule"
    serializer_class = WorkScheduleSerializer
    pagination_class = None
    permission_classes = [IsCatalogManager]

    def get_queryset(self):
        qs = WorkSchedule.objects.select_related("user", "counter")
        # Optional filters for the admin table.
        weekday = self.request.query_params.get("weekday")
        if weekday is not None and weekday != "":
            qs = qs.filter(weekday=weekday)
        return scope_to_hall(qs, self.request)

    def create(self, request, *args, **kwargs):
        data = request.data
        # Accept both the bulk shape and the legacy single-row shape.
        weekdays = data.get("weekdays")
        if weekdays is None and data.get("weekday") is not None:
            weekdays = [data["weekday"]]
        user_ids = data.get("user_ids")
        if user_ids is None and data.get("user_id") is not None:
            user_ids = [data["user_id"]]
        start_time = data.get("start_time")
        end_time = data.get("end_time")
        is_active = data.get("is_active", True)
        override_counter_id = data.get("counter_id")  # legacy explicit window

        errors = {}
        if not weekdays:
            errors["weekdays"] = ["Выберите хотя бы один день"]
        if not user_ids:
            errors["user_ids"] = ["Выберите хотя бы одного оператора"]
        if not start_time or not end_time:
            errors["start_time"] = ["Укажите время начала и конца"]
        elif str(start_time) >= str(end_time):
            errors["end_time"] = ["Конец смены должен быть позже начала"]
        if errors:
            raise ValidationError(errors)

        try:
            weekdays = sorted({int(w) for w in weekdays})
        except (TypeError, ValueError):
            raise ValidationError({"weekdays": ["Некорректные дни недели"]})
        if any(wd < 0 or wd > 6 for wd in weekdays):
            raise ValidationError({"weekdays": ["День недели вне диапазона 0–6"]})

        requester = request.user
        users = list(User.objects.filter(id__in=user_ids).select_related("counter"))
        created, updated, no_counter = [], 0, []
        with transaction.atomic():
            for user in users:
                counter = (
                    Counter.objects.filter(id=override_counter_id).first()
                    if override_counter_id
                    else user.counter
                )
                if counter is None:
                    no_counter.append(user.id)
                    continue
                # A hall_admin may only schedule onto counters in their own hall.
                if (
                    getattr(requester, "is_hall_admin", False)
                    and requester.hall_id
                    and counter.hall_id != requester.hall_id
                ):
                    raise PermissionDenied("Окно вне вашего зала")
                for wd in weekdays:
                    obj = WorkSchedule.objects.filter(
                        user=user, counter=counter, weekday=wd
                    ).first()
                    if obj is not None:
                        obj.start_time = start_time
                        obj.end_time = end_time
                        obj.is_active = is_active
                        obj.save(update_fields=["start_time", "end_time", "is_active", "updated_at"])
                        updated += 1
                        audit.log(request, "schedule.updated", target=obj.id)
                    else:
                        obj = WorkSchedule.objects.create(
                            user=user,
                            counter=counter,
                            weekday=wd,
                            start_time=start_time,
                            end_time=end_time,
                            is_active=is_active,
                        )
                        created.append(obj)
                        audit.log(request, "schedule.created", target=obj.id)

        return Response(
            {
                "created": WorkScheduleSerializer(created, many=True).data,
                "updated": updated,
                "no_counter": no_counter,
            },
            status=201,
        )


class ScheduleDetailView(AuditCRUDMixin, generics.RetrieveUpdateDestroyAPIView):
    audit_entity = "schedule"
    serializer_class = WorkScheduleSerializer
    permission_classes = [IsCatalogManager]

    def get_queryset(self):
        return scope_to_hall(
            WorkSchedule.objects.select_related("user", "counter"), self.request
        )


class ScheduleCurrentView(APIView):
    """GET /api/schedule/current → operators who are on duty right now per the
    plan (matches today's weekday + current local time). Powers the admin
    dashboard's "should be working" indicator. Public-read like the board."""

    permission_classes = [AllowAny]

    def get(self, request):
        now = timezone.localtime()
        weekday, t = now.weekday(), now.time()
        qs = scope_to_hall(
            WorkSchedule.objects.select_related("user", "counter").filter(
                is_active=True, weekday=weekday, start_time__lte=t, end_time__gt=t
            ),
            request,
        )
        return Response(WorkScheduleSerializer(qs, many=True).data)


# ---------- tickets (kiosk) ----------
class TicketCreateView(APIView):
    def post(self, request):
        ser = CreateTicketSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        category = ServiceCategory.objects.filter(id=data["category_id"]).first()
        if not category:
            return Response({"error": "unknown category"}, status=400)
        service = None
        if data.get("service_id"):
            service = Service.objects.filter(id=data["service_id"]).first()
        ticket = services.create_ticket(
            category=category, service=service, idempotency_key=data["idempotency_key"]
        )
        realtime.broadcast(
            [realtime.DISPLAY, realtime.OPERATORS, realtime.ADMIN], "ticket.created"
        )
        return Response(TicketSerializer(ticket).data, status=201)


# ---------- dashboard ----------
class DashboardView(APIView):
    def get(self, request):
        today = timezone.localdate()
        todays = Ticket.objects.filter(created_at__date=today)
        served_qs = todays.filter(status=TicketStatus.SERVED)
        wait = (
            todays.filter(called_at__isnull=False)
            .annotate(w=F("called_at") - F("created_at"))
            .aggregate(avg=Avg("w"))["avg"]
        )
        avg_wait_min = int(wait.total_seconds() // 60) if wait else 0

        hourly = []
        for hour in range(8, 18):
            hourly.append(
                {
                    "hour": hour,
                    "issued": todays.filter(created_at__hour=hour).count(),
                    "served": served_qs.filter(called_at__hour=hour).count(),
                }
            )

        recent = []
        for t in Ticket.objects.select_related("category", "service", "counter").order_by(
            "-created_at"
        )[:10]:
            recent.append(
                {
                    "id": str(t.id),
                    "number": t.number,
                    "category_code": t.category.code,
                    "service": {
                        "name_kaa": t.service.name_kaa,
                        "name_ru": t.service.name_ru,
                        "name_uz": t.service.name_uz,
                        "name_en": t.service.name_en,
                    } if t.service_id else None,
                    "status": t.status,
                    "counter_number": t.counter.number if t.counter_id else None,
                    "issued_at": t.created_at.isoformat(),
                }
            )

        return Response(
            {
                "metrics": {
                    "ticketsToday": todays.count(),
                    "avgWaitMinutes": avg_wait_min,
                    "activeCounters": Counter.objects.filter(is_active=True).count(),
                    "served": served_qs.count(),
                },
                "hourly": hourly,
                "recent": recent,
            }
        )


# ---------- statistics ----------
def _stats_qs(request):
    params = request.query_params
    qs = Ticket.objects.all()
    if params.get("hall"):
        qs = qs.filter(hall_id=params["hall"])
    if params.get("from"):
        qs = qs.filter(created_at__date__gte=params["from"])
    if params.get("to"):
        qs = qs.filter(created_at__date__lte=params["to"])
    return scope_to_hall(qs, request)  # hall_admin sees only their hall


def _compute_stats(qs):
    served = qs.filter(status=TicketStatus.SERVED)
    wait = (
        qs.filter(called_at__isnull=False)
        .annotate(w=F("called_at") - F("created_at"))
        .aggregate(a=Avg("w"))["a"]
    )
    svc = (
        served.filter(finished_at__isnull=False, called_at__isnull=False)
        .annotate(s=F("finished_at") - F("called_at"))
        .aggregate(a=Avg("s"))["a"]
    )
    hourly = []
    peak_hour, peak_count = None, -1
    for hour in range(8, 19):
        cnt = qs.filter(created_at__hour=hour).count()
        hourly.append({"hour": hour, "issued": cnt})
        if cnt > peak_count:
            peak_hour, peak_count = hour, cnt
    return {
        "issued": qs.count(),
        "served": served.count(),
        "skipped": qs.filter(status=TicketStatus.SKIPPED).count(),
        "avg_wait_minutes": int(wait.total_seconds() // 60) if wait else 0,
        "avg_service_minutes": int(svc.total_seconds() // 60) if svc else 0,
        "peak_hour": peak_hour if peak_count > 0 else None,
        "hourly": hourly,
    }


class StatsView(APIView):
    """Aggregated stats: served / avg wait / avg service / skipped / peak hour.
    Filters: ?hall=&from=YYYY-MM-DD&to=YYYY-MM-DD."""

    def get(self, request):
        return Response(_compute_stats(_stats_qs(request)))


class StatsExportView(APIView):
    """Excel (.xlsx): 1-varaq — talonlar ro'yxati (avtofiltr), 2-varaq —
    operatorlar bo'yicha jamlanma (kim nechta odam qabul qildi, jami necha
    daqiqa ishladi, bir odamga o'rtacha necha daqiqa). Saralash — Excel ichida."""

    # Talon holati — o'zbekcha (admin paneldagi bilan bir xil).
    STATUS_UZ = {
        "waiting": "Kutilmoqda",
        "called": "Chaqirilgan",
        "serving": "Xizmat ko'rsatilmoqda",
        "served": "Xizmat ko'rsatilgan",
        "skipped": "O'tkazib yuborilgan",
        "cancelled": "Bekor qilingan",
    }

    HEADERS = [
        "Raqam", "Sana", "Zal", "Kategoriya", "Xizmat", "Oyna", "Operator",
        "Holat", "Berilgan", "Chaqirilgan", "Yakunlangan",
        "Kutish (daq)", "Xizmat vaqti (daq)",
    ]
    WIDTHS = [9, 11, 18, 26, 38, 7, 28, 22, 10, 11, 12, 12, 16]

    # 2-varaq: operatorlar bo'yicha jamlanma. 1-ustun — hisobot davri (Sana).
    OP_HEADERS = [
        "Sana", "Operator", "Qabul qilingan (kishi)", "Jami xizmat (daq)",
        "O'rtacha (daq/kishi)", "O'tkazib yuborilgan",
    ]
    OP_WIDTHS = [20, 30, 22, 18, 22, 22]

    # 3-varaq: xizmatlar bo'yicha (eng ko'p so'ralgan xizmatlar yuqorida).
    SVC_HEADERS = [
        "Sana", "Xizmat", "Kategoriya", "So'ralgan (marta)", "Xizmat ko'rsatilgan",
        "O'rtacha (daq)",
    ]
    SVC_WIDTHS = [20, 42, 12, 18, 20, 14]

    def get(self, request):
        from collections import defaultdict
        from io import BytesIO

        from django.http import HttpResponse
        from django.utils.timezone import localtime
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        qs = (
            _stats_qs(request)
            .select_related("hall", "category", "service", "counter", "operator")
            .order_by("created_at")  # neytral tartib; saralash — avtofiltr orqali
        )

        def fdate(dt):
            return localtime(dt).strftime("%d.%m.%Y") if dt else ""

        def ftime(dt):
            return localtime(dt).strftime("%H:%M") if dt else ""

        def fmins(a, b):  # son sifatida (Excel raqamli saralashi uchun)
            return round((b - a).total_seconds() / 60) if a and b else None

        def uz(obj):
            return (obj.name_uz or obj.name_ru or "") if obj else ""

        # Hisobot davri — agregat varaqlardagi (Operatorlar/Xizmatlar) "Sana"
        # ustuni: tanlangan davr yoki butun tarix.
        def fday(s):  # "YYYY-MM-DD" → "DD.MM.YYYY"
            try:
                y, m, d = s.split("-")
                return f"{d}.{m}.{y}"
            except (ValueError, AttributeError):
                return s or ""
        _df = request.query_params.get("from")
        _dt = request.query_params.get("to")
        if _df and _dt:
            period = f"{fday(_df)}–{fday(_dt)}"
        elif _df:
            period = f"{fday(_df)}–…"
        elif _dt:
            period = f"…–{fday(_dt)}"
        else:
            period = "Hammasi"

        head_font = Font(bold=True, color="FFFFFF")
        head_fill = PatternFill("solid", fgColor="2563EB")

        def finalize(ws, headers, widths):
            """Sarlavha stili + avtofiltr + muzlatilgan sarlavha + kengliklar."""
            for cell in ws[1]:
                cell.font = head_font
                cell.fill = head_fill
                cell.alignment = Alignment(vertical="center")
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
            ws.freeze_panes = "A2"
            for i, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width

        wb = Workbook()

        # ── 1-varaq: talonlar ──
        ws = wb.active
        ws.title = "Talonlar"
        ws.append(self.HEADERS)

        # operatorlar bo'yicha jamlanma (2-varaq uchun yig'amiz)
        agg = defaultdict(lambda: {"served": 0, "service_min": 0, "skipped": 0})
        # xizmatlar bo'yicha jamlanma (3-varaq uchun)
        svc_agg = defaultdict(
            lambda: {"category": "", "requested": 0, "served": 0,
                     "service_sum": 0, "served_timed": 0}
        )

        for t in qs:
            operator = ""
            if t.operator_id:
                operator = (
                    t.operator.name
                    or t.operator.get_full_name()
                    or t.operator.username
                )
            category = ""
            if t.category_id:
                cat_name = uz(t.category)
                category = (
                    f"{t.category.code} · {cat_name}" if cat_name else t.category.code
                )
            service_min = fmins(t.called_at, t.finished_at)
            ws.append([
                t.number,
                fdate(t.created_at),
                uz(t.hall) if t.hall_id else "",
                category,
                uz(t.service) if t.service_id else "",
                t.counter.number if t.counter_id else "",
                operator,
                self.STATUS_UZ.get(t.status, t.status),
                ftime(t.created_at),
                ftime(t.called_at),
                ftime(t.finished_at),
                fmins(t.created_at, t.called_at),
                service_min,
            ])
            if operator:
                if t.status == "served":
                    agg[operator]["served"] += 1
                    if service_min:
                        agg[operator]["service_min"] += service_min
                elif t.status == "skipped":
                    agg[operator]["skipped"] += 1

            if t.service_id:
                s = svc_agg[uz(t.service)]
                s["category"] = t.category.code if t.category_id else ""
                s["requested"] += 1
                if t.status == "served":
                    s["served"] += 1
                    if service_min:
                        s["service_sum"] += service_min
                        s["served_timed"] += 1

        finalize(ws, self.HEADERS, self.WIDTHS)

        # ── 2-varaq: operatorlar bo'yicha jamlanma (ВСЕ операторы, даже без талонов) ──
        from django.contrib.auth import get_user_model

        ws2 = wb.create_sheet("Operatorlar")
        ws2.append(self.OP_HEADERS)
        # Все операторы по справочнику + те, кто реально обслуживал (admin/др. роли).
        # scope_to_hall: chief видит всех, hall_admin — только операторов своего зала
        # (talonlar/xizmatlar листы уже скоуплены через qs).
        op_qs = scope_to_hall(
            get_user_model().objects.filter(role="operator"), request
        )
        names, seen = [], set()
        for u in op_qs.order_by("name", "username"):
            nm = u.name or u.get_full_name() or u.username
            if nm and nm not in seen:
                seen.add(nm); names.append(nm)
        for nm in agg:
            if nm not in seen:
                seen.add(nm); names.append(nm)
        for op in sorted(names):
            a = agg.get(op, {"served": 0, "service_min": 0, "skipped": 0})
            avg = round(a["service_min"] / a["served"]) if a["served"] else None
            ws2.append([period, op, a["served"], a["service_min"], avg, a["skipped"]])
        finalize(ws2, self.OP_HEADERS, self.OP_WIDTHS)

        # ── 3-varaq: xizmatlar (eng ko'p so'ralgani yuqorida) ──
        ws3 = wb.create_sheet("Xizmatlar")
        ws3.append(self.SVC_HEADERS)
        for sname in sorted(svc_agg, key=lambda n: -svc_agg[n]["requested"]):
            s = svc_agg[sname]
            avg = (
                round(s["service_sum"] / s["served_timed"])
                if s["served_timed"] else None
            )
            ws3.append([period, sname, s["category"], s["requested"], s["served"], avg])
        finalize(ws3, self.SVC_HEADERS, self.SVC_WIDTHS)

        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        resp["Content-Disposition"] = 'attachment; filename="navbat-statistika.xlsx"'
        return resp


# ---------- operator sessions ----------
class OperatorSessionCreateView(APIView):
    def post(self, request):
        counter = Counter.objects.filter(id=request.data.get("counter_id")).first()
        if not counter:
            return Response({"error": "unknown counter"}, status=404)
        session = OperatorSession.objects.create(
            user_id=request.data.get("user_id"), counter=counter
        )
        return Response(OperatorSessionSerializer(session).data, status=201)


class OperatorSessionDetailView(APIView):
    def patch(self, request, pk):
        session = OperatorSession.objects.filter(id=pk).first()
        if not session:
            return Response({"error": "not found"}, status=404)
        services.set_session_status(session, request.data.get("status"))
        return Response(OperatorSessionSerializer(session).data)


# ---------- queue + current ----------
class QueueView(APIView):
    def get(self, request):
        counter = Counter.objects.filter(
            id=request.query_params.get("counter_id")
        ).first()
        if not counter:
            return Response({"error": "unknown counter"}, status=404)
        q = services.queue_for_counter(counter)[:20]
        return Response(TicketSerializer(q, many=True).data)


class CurrentTicketView(APIView):
    def get(self, request):
        counter = Counter.objects.filter(
            id=request.query_params.get("counter_id")
        ).first()
        current = services.current_for_counter(counter) if counter else None
        if not current:
            # JSON `null` (DRF Response(None) yields an empty body, which the
            # frontend's res.json() can't parse).
            return JsonResponse(None, safe=False)
        return Response(TicketSerializer(current).data)


# ---------- transitions ----------
class CallNextView(APIView):
    def post(self, request):
        counter = Counter.objects.filter(id=request.data.get("counter_id")).first()
        if not counter:
            return Response({"error": "unknown counter"}, status=404)
        operator_id = request.data.get("operator_id")
        # Optional ticket_id → operator picked a SPECIFIC client from the queue;
        # otherwise fall back to calling the oldest eligible one (the big button).
        ticket_id = request.data.get("ticket_id")
        if ticket_id:
            ticket = services.call_specific(
                counter=counter, ticket_id=ticket_id, operator_id=operator_id
            )
            if not ticket:
                return Response({"error": "ticket not callable"}, status=409)
        else:
            ticket = services.call_next(counter=counter, operator_id=operator_id)
            if not ticket:
                return Response({"error": "queue empty"}, status=409)
        audit.log(
            request, "ticket.called", target=ticket.number,
            actor_label=str(request.data.get("operator_id") or ""),
            counter=counter.number,
        )
        realtime.broadcast(
            [realtime.DISPLAY, realtime.OPERATORS, realtime.ADMIN], "ticket.called"
        )
        return Response(TicketSerializer(ticket).data)


class TicketActionView(APIView):
    """finish / skip via the URL-configured `action`."""

    action = None

    def post(self, request, pk):
        ticket = Ticket.objects.filter(id=pk).first()
        if not ticket:
            return Response({"error": "not found"}, status=404)
        if self.action == "finish":
            services.finish(ticket)
        elif self.action == "skip":
            services.skip(ticket)
        audit.log(request, f"ticket.{self.action}ed", target=ticket.number)
        realtime.broadcast(
            [realtime.DISPLAY, realtime.OPERATORS, realtime.ADMIN],
            f"ticket.{self.action}ed",
        )
        return Response(TicketSerializer(ticket).data)


class TicketRecallView(APIView):
    """Repeat the announcement for a current call (operator 'Repeat' button)."""

    def post(self, request, pk):
        ticket = Ticket.objects.filter(id=pk).first()
        if not ticket:
            return Response({"error": "not found"}, status=404)
        services.recall(ticket)
        audit.log(request, "ticket.recalled", target=ticket.number)
        realtime.broadcast(
            [realtime.DISPLAY, realtime.OPERATORS, realtime.ADMIN], "ticket.called"
        )
        return Response(TicketSerializer(ticket).data)


class TicketTransferView(APIView):
    def post(self, request, pk):
        ticket = Ticket.objects.filter(id=pk).first()
        if not ticket:
            return Response({"error": "not found"}, status=404)
        counter = Counter.objects.filter(id=request.data.get("counter_id")).first()
        if not counter:
            return Response({"error": "unknown counter"}, status=404)
        services.transfer(ticket, counter)
        audit.log(
            request, "ticket.transferred", target=ticket.number,
            to_counter=counter.number,
        )
        realtime.broadcast(
            [realtime.DISPLAY, realtime.OPERATORS, realtime.ADMIN], "ticket.transferred"
        )
        return Response(TicketSerializer(ticket).data)


# ---------- display ----------
class DisplayActiveView(APIView):
    def get(self, request):
        hall_id = request.query_params.get("hall_id")
        calls = services.active_calls(limit=12, hall_id=hall_id)
        return Response(DisplayCallSerializer(calls, many=True).data)


class DisplayBoardView(APIView):
    """All active windows + the current call on each (null when idle).
    Powers the board's window strip — scoped to a hall when ?hall_id= given."""

    def get(self, request):
        counters = Counter.objects.filter(is_active=True)
        hall_id = request.query_params.get("hall_id")
        if hall_id:
            counters = counters.filter(hall_id=hall_id)
        counters = list(counters)
        for c in counters:
            c.current = services.current_for_counter(c)
        return Response(DisplayBoardCounterSerializer(counters, many=True).data)


class DisplayWaitingView(APIView):
    """The waiting queue (issued, not yet called) shown on the board so a
    visitor sees their number before it's called. Scoped to a hall when given."""

    def get(self, request):
        hall_id = request.query_params.get("hall_id")
        return Response(
            DisplayWaitingSerializer(
                services.waiting_list(20, hall_id=hall_id), many=True
            ).data
        )


# ---------- sync (local-first) ----------
class SyncCatalogView(APIView):
    """cloud → local: full catalog snapshot the local box mirrors."""

    permission_classes = [HasSyncToken]

    def get(self, request):
        return Response(sync.catalog_snapshot())


class SyncEventsView(APIView):
    """local → cloud: upsert tickets / sessions / audit pushed from a box."""

    permission_classes = [HasSyncToken]

    def post(self, request):
        counts = sync.ingest_events(request.data or {})
        return Response({"ok": True, **counts})


# ---------- audit ----------
class AuditListView(generics.ListAPIView):
    """Audit journal with optional filters (?action=&actor=&from=&to=). Capped
    at 200 newest. Chief-only once auth enforcement lands."""

    serializer_class = AuditLogSerializer
    pagination_class = None
    permission_classes = [IsChief]

    def get_queryset(self):
        qs = AuditLog.objects.select_related("actor").all()
        p = self.request.query_params
        if p.get("action"):
            qs = qs.filter(action=p["action"])
        if p.get("actor"):
            qs = qs.filter(actor_id=p["actor"])
        if p.get("from"):
            qs = qs.filter(created_at__gte=p["from"])
        if p.get("to"):
            qs = qs.filter(created_at__lte=p["to"])
        return qs[:200]


class AuditExportView(APIView):
    """Excel (.xlsx) audit jurnali. Filtrlar: ?action=&from=&to= (kun bo'yicha).
    Avtofiltr + o'zbekcha sarlavhalar, lola-stilidagi statistika eksporti kabi.

    Permission: StatsExportView kabi default (AllowAny) — yuklab olish <a href>
    orqali bo'lgani uchun brauzer JWT-sarlavha yubormaydi (aks holda 401)."""

    HEADERS = ["Sana", "Vaqt", "Amal", "Obyekt", "Kim", "Tafsilotlar"]
    WIDTHS = [12, 9, 26, 22, 30, 48]
    ACTION_UZ = {
        "ticket.called": "Chaqiruv",
        "ticket.recalled": "Qayta chaqiruv",
        "ticket.finished": "Yakunlash",
        "ticket.skipped": "O'tkazib yuborish",
        "ticket.transferred": "Boshqa oynaga",
        "counter.created": "Oyna yaratildi",
        "counter.updated": "Oyna o'zgartirildi",
        "counter.deleted": "Oyna o'chirildi",
        "queue.reset": "Navbat tiklandi",
        "auth.login": "Kirish",
    }

    @staticmethod
    def _fmt_meta(meta):
        """meta-JSON → o'qiladigan matn: {'counter':'9'} → 'Oyna 9'."""
        if not meta:
            return ""
        parts = []
        for k, v in meta.items():
            if k == "counter":
                parts.append(f"Oyna {v}")
            elif k == "cancelled":
                parts.append(f"{v} ta bekor qilingan")
            else:
                parts.append(f"{k}: {v}")
        return ", ".join(parts)

    def get(self, request):
        from io import BytesIO

        from django.http import HttpResponse
        from django.utils.timezone import localtime
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        qs = AuditLog.objects.select_related("actor").all()
        p = request.query_params
        if p.get("action"):
            qs = qs.filter(action=p["action"])
        if p.get("from"):
            qs = qs.filter(created_at__gte=p["from"])
        if p.get("to"):
            qs = qs.filter(created_at__lte=p["to"])
        qs = qs.order_by("-created_at")[:5000]

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit"
        ws.append(self.HEADERS)
        head_font = Font(bold=True, color="FFFFFF")
        head_fill = PatternFill("solid", fgColor="2563EB")
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")

        for a in qs:
            actor = (a.actor.name or a.actor.get_full_name() or a.actor.username) if a.actor_id else a.actor_label
            ws.append([
                localtime(a.created_at).strftime("%d.%m.%Y"),
                localtime(a.created_at).strftime("%H:%M"),
                self.ACTION_UZ.get(a.action, a.action),
                a.target or "",
                actor or "",
                self._fmt_meta(a.meta),
            ])

        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.HEADERS))}{ws.max_row}"
        ws.freeze_panes = "A2"
        for i, width in enumerate(self.WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        resp["Content-Disposition"] = 'attachment; filename="audit.xlsx"'
        return resp


class DisplaySettingsView(APIView):
    """Board config (YouTube URL). GET public (the board reads it), PATCH chief."""

    def get_permissions(self):
        return [AllowAny()] if self.request.method == "GET" else [IsChief()]

    def get(self, request):
        return Response(DisplaySettingsSerializer(DisplaySettings.load()).data)

    def patch(self, request):
        ser = DisplaySettingsSerializer(
            DisplaySettings.load(), data=request.data, partial=True
        )
        ser.is_valid(raise_exception=True)
        ser.save()
        # tell the board to reload its media zone live
        realtime.broadcast([realtime.DISPLAY], "display.settings")
        return Response(ser.data)


class VoiceClipListCreateView(APIView):
    """Загружаемая озвучка вызовов. GET public (табло + админка читают),
    POST chief — upsert клипа по (kind, key) с заменой файла."""

    parser_classes = [MultiPartParser, FormParser]

    def get_permissions(self):
        return [AllowAny()] if self.request.method == "GET" else [IsChief()]

    def get(self, request):
        clips = VoiceClip.objects.all()
        return Response(VoiceClipSerializer(clips, many=True).data)

    def post(self, request):
        ser = VoiceClipSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        kind = ser.validated_data["kind"]
        key = ser.validated_data["key"]
        new_file = ser.validated_data["file"]
        clip = VoiceClip.objects.filter(kind=kind, key=key).first()
        if clip:
            clip.file.delete(save=False)  # убрать старый файл → имя переиспользуется
            clip.file = new_file
            clip.enabled = True  # загрузка = «использовать свою запись»
            clip.save()
        else:
            clip = VoiceClip.objects.create(kind=kind, key=key, file=new_file)
        realtime.broadcast([realtime.DISPLAY], "display.voice")
        return Response(VoiceClipSerializer(clip).data, status=201)


class VoiceClipDetailView(APIView):
    """PATCH — переключить enabled (lola ⇄ своя запись). DELETE — удалить клип
    (слот возвращается к встроенному lola). Обе операции — chief."""

    permission_classes = [IsChief]

    def patch(self, request, pk):
        clip = VoiceClip.objects.filter(pk=pk).first()
        if not clip:
            return Response({"error": "not found"}, status=404)
        enabled = request.data.get("enabled")
        if not isinstance(enabled, bool):
            # строго bool: bool("false") == True, поэтому строки не принимаем
            return Response({"error": "enabled must be boolean"}, status=400)
        clip.enabled = enabled
        clip.save(update_fields=["enabled", "updated_at"])
        realtime.broadcast([realtime.DISPLAY], "display.voice")
        return Response(VoiceClipSerializer(clip).data)

    def delete(self, request, pk):
        clip = VoiceClip.objects.filter(pk=pk).first()
        if not clip:
            return Response({"error": "not found"}, status=404)
        clip.file.delete(save=False)
        clip.delete()
        realtime.broadcast([realtime.DISPLAY], "display.voice")
        return Response(status=204)
